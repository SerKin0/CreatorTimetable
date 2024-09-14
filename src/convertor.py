import json
import os.path

from openpyxl.reader.excel import load_workbook


class Convertor:
    def __init__(self):
        self.timetable = {}

    def xlsx2json(self, path_excel: str, path_json: str):
        if not os.path.exists(path_excel):
            return False
        workbook = load_workbook(path_excel)
        for day in workbook.get_sheet_names():
            self.timetable[day] = []
            table = workbook[day]
            for number in range(1, table.max_row - 1):
                temp = {"number": number}
                if (table.cell(2 + number, 2).value is None) and (table.cell(2 + number, 7).value is None):
                    temp["free"] = False
                    self.timetable[day].append(temp)
                    continue
                temp["free"] = True
                temp["left"] = {
                    "summary": table.cell(2 + number, 2).value,
                    "location": table.cell(2 + number, 3).value,
                    "description": table.cell(2 + number, 4).value,
                    "start": table.cell(2 + number, 5).value,
                    "end": table.cell(2 + number, 6).value,
                }
                temp["right"] = {
                    "summary": table.cell(2 + number, 7).value,
                    "location": table.cell(2 + number, 8).value,
                    "description": table.cell(2 + number, 9).value,
                    "start": table.cell(2 + number, 10).value,
                    "end": table.cell(2 + number, 11).value,
                }
                self.timetable[day].append(temp)

        # Создание json файла
        json_string = json.dumps(self.timetable)
        with open(path_json, "w", encoding="utf-8") as file:
            file.write(json_string)
        return True
